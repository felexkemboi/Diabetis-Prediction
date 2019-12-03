

import numpy as np
from flask import Flask, request, jsonify,render_template
import pickle


from flaskext.mysql import MySQL


app=Flask(__name__)
model = pickle.load(open('model.pkl','rb'))

@app.route('/',methods = ['POST','GET'])
def home():
    return render_template("home.html")


def ValuePredictor(to_predict_list):
    to_predict = np.array(to_predict_list).reshape(1,8)
    loaded_model = pickle.load(open("model.pkl","rb"))
    result = loaded_model.predict(to_predict)
    return result[0]

@app.route('/predict',methods = ['POST','GET'])
def index():
    if request.method == 'POST':
        to_predict_list = request.form.to_dict()
        to_predict_list=list(to_predict_list.values())

        print(to_predict_list)

        to_predict_list = list(map(float, to_predict_list))
        print("yeees")

        print(to_predict_list)

        result = ValuePredictor(to_predict_list)
        print(result)

        if int(result)==1:
            prediction='You are Diabetic'
        else:
            prediction='You are Not Diabetic'
        print(prediction)
        return render_template("index.html",prediction=prediction)
    else:
        return render_template("index.html")



if __name__ == '__main__':
    app.run(port=5000, debug=True)

"""
from selenium import webdriver
from underscore import _
from openpyxl import load_workbook, cell
from selenium.webdriver.support.ui import WebDriverWait
import openpyxl
from selenium.webdriver.support import expected_conditions  as EC
from selenium.webdriver.common.by import By
import time
import selenium.webdriver as webdriver
import selenium.webdriver.support.ui as ui
from selenium.webdriver.common.keys import Keys
from time import sleep
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
driver = webdriver.Chrome(ChromeDriverManager().install())
driver.get("https://www.coursehero.com/subjects/political-science/questions")
wait = WebDriverWait(driver, 3)
linkindex = 0
previous = "Showing 1 to 8 of 3,821"
j = 0  
row = 0 
"""


#to tell flask what url shoud trigger the function index()
"""
@app.route('/')
def index():
    return render_template('index.html',model=model)
    """



"""

@app.route('/test',methods = ['POST','GET'])
def index1():
    while True:
        allquizes = driver.find_elements_by_class_name("taxonomy-section-heading")
        now = str(allquizes[0].text)

        if previous == now:
            global row
            global j
            while j <= 3821:
                allquizes = driver.find_elements_by_class_name("taxonomy-section-heading")
                now = str(allquizes[0].text)
                print()
                print("Political science" + str(now))
                wb = openpyxl.load_workbook('excel.xlsx')
                sheet = wb.get_sheet_by_name("test")
                sheet["A1"] = "Title"
                sheet["B1"] = "Question"


                link1 = driver.find_elements_by_tag_name('h6')[0].find_element_by_tag_name('a') #.get_attribute('href')
                main_window = driver.current_window_handle
                wait
                time.sleep(1)
                link1.send_keys(Keys.CONTROL + Keys.RETURN)
                driver.switch_to.window(driver.window_handles[1])
                WebDriverWait(driver, 2)
                questions =  driver.find_elements_by_xpath('//*[@id="qal_thread-item_question"]/div/div[1]/span/p[4]')
                title  = driver.find_elements_by_xpath('//*[@id="qal_thread-item_question"]/div/div[1]/span/h1')
                print("Add the  Question number " + str(j))
                print("The last row is  " + str(row))
                if len(title)> 0:
                    #sheet["A" + str(row)].value = title[0].text
                    sheet.cell(row=row, column=1).value = title[0].text

                else:
                    #print("No title")
                    pass
                if len(questions)> 0:
                    #sheet["B" + str(row)].value = questions[0].text
                    sheet.cell(row=row, column=2).value = questions[0].text

                else:
                    #print("No questions just attachments")
                    pass

                j = j+1
                row = row + 1
                print()
                driver.close()
                driver.switch_to.window(main_window)
                #del link1

                link2 = driver.find_elements_by_tag_name('h6')[1].find_element_by_tag_name('a') #.get_attribute('href')
                main_window = driver.current_window_handle
                time.sleep(1)
                link2.send_keys(Keys.CONTROL + Keys.RETURN)
                driver.switch_to.window(driver.window_handles[1])
                WebDriverWait(driver, 2)
                questions =  driver.find_elements_by_xpath('//*[@id="qal_thread-item_question"]/div/div[1]/span/p[4]')
                title  = driver.find_elements_by_xpath('//*[@id="qal_thread-item_question"]/div/div[1]/span/h1')
                #WebDriverWait(driver, 6)
                print("Add the  Question number " + str(j-1))
                print("The last row is  " + str(row))
                if len(title)> 0:
                    sheet.cell(row=row, column=1).value = title[0].text
                else:
                    #print("No title")
                    pass
                if len(questions)> 0:
                    sheet.cell(row=row, column=2).value = questions[0].text
                else:
                    #print("No questions just attachments")
                    pass
                j = j+1
                row = row + 1
                print()
                driver.close()
                driver.switch_to.window(main_window)




                link3 = driver.find_elements_by_tag_name('h6')[2].find_element_by_tag_name('a') #.get_attribute('href')
                main_window = driver.current_window_handle
                time.sleep(1)
                link3.send_keys(Keys.CONTROL + Keys.RETURN)
                driver.switch_to.window(driver.window_handles[1])
                WebDriverWait(driver, 2)
                questions =  driver.find_elements_by_xpath('//*[@id="qal_thread-item_question"]/div/div[1]/span/p[4]')
                title  = driver.find_elements_by_xpath('//*[@id="qal_thread-item_question"]/div/div[1]/span/h1')
                #WebDriverWait(driver, 6)
                print("Add the  Question number " + str(j-1))
                print("The last row is  " + str(row))
                if len(title)> 0:
                    sheet.cell(row=row, column=1).value = title[0].text
                else:
                    #print("No title")
                    pass
                if len(questions)> 0:
                    sheet.cell(row=row, column=2).value = questions[0].text
                else:
                    #print("No questions just attachments")
                    pass
                j = j+1
                row = row + 1
                print()
                driver.close()
                driver.switch_to.window(main_window)



                link4 = driver.find_elements_by_tag_name('h6')[3].find_element_by_tag_name('a') #.get_attribute('href')
                main_window = driver.current_window_handle
                time.sleep(1)
                link4.send_keys(Keys.CONTROL + Keys.RETURN)
                driver.switch_to.window(driver.window_handles[1])
                WebDriverWait(driver, 2)
                questions =  driver.find_elements_by_xpath('//*[@id="qal_thread-item_question"]/div/div[1]/span/p[4]')
                title  = driver.find_elements_by_xpath('//*[@id="qal_thread-item_question"]/div/div[1]/span/h1')
                #WebDriverWait(driver, 6)
                print("Add the  Question number " + str(j))
                print("The last row is  " + str(row))
                if len(title)> 0:
                    sheet.cell(row=row, column=1).value = title[0].text
                else:
                    #print("No title")
                    pass
                if len(questions)> 0:
                    sheet.cell(row=row, column=2).value = questions[0].text
                else:
                    #print("No questions just attachments")
                    pass
                j = j+1
                row = row + 1
                print()
                driver.close()
                driver.switch_to.window(main_window)


                link5 = driver.find_elements_by_tag_name('h6')[4].find_element_by_tag_name('a') #.get_attribute('href')
                main_window = driver.current_window_handle
                time.sleep(1)
                link5.send_keys(Keys.CONTROL + Keys.RETURN)
                driver.switch_to.window(driver.window_handles[1])
                WebDriverWait(driver, 2)
                questions =  driver.find_elements_by_xpath('//*[@id="qal_thread-item_question"]/div/div[1]/span/p[4]')
                title  = driver.find_elements_by_xpath('//*[@id="qal_thread-item_question"]/div/div[1]/span/h1')
                #WebDriverWait(driver, 6)
                print("Add the  Question number " + str(j))
                print("The last row is  " + str(row))
                if len(title)> 0:
                    sheet.cell(row=row, column=1).value = title[0].text
                else:
                    #print("No title")
                    pass
                if len(questions)> 0:
                    sheet.cell(row=row, column=2).value = questions[0].text
                else:
                    #print("No questions just attachments")
                    pass
                j = j+1
                row = row + 1
                print()
                driver.close()
                driver.switch_to.window(main_window)



                link6 = driver.find_elements_by_tag_name('h6')[5].find_element_by_tag_name('a') #.get_attribute('href')
                main_window = driver.current_window_handle
                time.sleep(1)
                link6.send_keys(Keys.CONTROL + Keys.RETURN)
                driver.switch_to.window(driver.window_handles[1])
                WebDriverWait(driver, 2)
                questions =  driver.find_elements_by_xpath('//*[@id="qal_thread-item_question"]/div/div[1]/span/p[4]')
                title  = driver.find_elements_by_xpath('//*[@id="qal_thread-item_question"]/div/div[1]/span/h1')
                #WebDriverWait(driver, 6)
                print("Add the  Question number " + str(j))
                print("The last row is  " + str(row))
                if len(title)> 0:
                    sheet.cell(row=row, column=1).value = title[0].text
                else:
                    #print("No title")
                    pass
                if len(questions)> 0:
                    sheet.cell(row=row, column=2).value = questions[0].text
                else:
                    #print("No questions just attachments")
                    pass
                j = j+1
                row = row + 1
                print()
                j = j+1
                driver.close()
                driver.switch_to.window(main_window)


                link7 = driver.find_elements_by_tag_name('h6')[6].find_element_by_tag_name('a') #.get_attribute('href')
                main_window = driver.current_window_handle
                time.sleep(1)
                link7.send_keys(Keys.CONTROL + Keys.RETURN)
                driver.switch_to.window(driver.window_handles[1])
                WebDriverWait(driver, 2)
                questions =  driver.find_elements_by_xpath('//*[@id="qal_thread-item_question"]/div/div[1]/span/p[4]')
                title  = driver.find_elements_by_xpath('//*[@id="qal_thread-item_question"]/div/div[1]/span/h1')
                #WebDriverWait(driver, 6)
                print("Add the  Question number " + str(j))
                print("The last row is  " + str(row))
                if len(title)> 0:
                    sheet.cell(row=row, column=1).value = title[0].text
                else:
                    #print("No title")
                    pass
                if len(questions)> 0:
                    sheet.cell(row=row, column=2).value = questions[0].text
                else:
                    #print("No questions just attachments")
                    pass
                j = j+1
                row = row + 1
                print()
                driver.close()
                driver.switch_to.window(main_window)


                link8 = driver.find_elements_by_tag_name('h6')[7].find_element_by_tag_name('a') #.get_attribute('href')
                main_window = driver.current_window_handle
                time.sleep(2)
                link8.send_keys(Keys.CONTROL + Keys.RETURN)
                driver.switch_to.window(driver.window_handles[1])
                WebDriverWait(driver, 2)
                questions =  driver.find_elements_by_xpath('//*[@id="qal_thread-item_question"]/div/div[1]/span/p[4]')
                title  = driver.find_elements_by_xpath('//*[@id="qal_thread-item_question"]/div/div[1]/span/h1')
                #WebDriverWait(driver, 6)
                print("Add the  Question number " + str(j))
                print("The last row is  " + str(row))
                if len(title)> 0:
                    sheet.cell(row=row, column=1).value = title[0].text
                else:
                    #print("No title")
                    pass
                if len(questions)> 0:
                    sheet.cell(row=row, column=2).value = questions[0].text
                else:
                    pass
                j = j+1
                row = row + 1
                print()

                driver.close()
                driver.switch_to.window(main_window)
                wb.save("excel.xlsx")


                element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="pagination-next"]')))
                element = driver.find_element_by_xpath('//*[@id="pagination-next"]') #.click()
                driver.execute_script("arguments[0].click();", element)
                time.sleep(3)
            allquizes = driver.find_elements_by_class_name("taxonomy-section-heading")
            now = str(allquizes[0].text)
            print(now)
            print("The last question is  " + str(j))
            
            
            driver.quit()
            break
        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="pagination-next"]')))
        element = driver.find_element_by_xpath('//*[@id="pagination-next"]') #.click()
        driver.execute_script("arguments[0].click();", element)
        time.sleep(3)

        if j == 3821:
            break
    return True
"""
